#!/usr/bin/env python
import os
import sys

if __name__ == "__main__":
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "django_project.settings")

    from django.core.management import execute_from_command_line

    execute_from_command_line(sys.argv)



import string
from random import *
import openpyxl
from django.contrib.auth.models import User
from app.models import Iscrizione
characters =  string.digits


utenti= User.objects.all()
a=1
for utenti in utenti:
	print(a)
	print(utenti)
	Iscrizione.objects.create(user=utenti)
	a=a+1



# riga = 270
# while riga < 1443:
#         riga = riga + 1
#         excel_document = openpyxl.load_workbook('listastudenti.xlsx')
#         sheet = excel_document.get_sheet_by_name('Foglio1')
#         nome_utente=(sheet.cell(row = riga, column = 1).value)
#         classe= (sheet.cell(row = riga, column = 2).value)
#         utente=str(nome_utente)+str(classe)
#         print (riga)
#         utente = User.objects.get(username=utente)
#         Iscrizione.objects.create(user=utente)