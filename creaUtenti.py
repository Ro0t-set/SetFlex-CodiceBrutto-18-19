#!/usr/bin/env python
import os
import sys

if __name__ == "__main__":
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "django_project.settings")

    from django.core.management import execute_from_command_line

    execute_from_command_line(sys.argv)

import math
import string
from django.contrib.auth.models import User


#########rddfhgjkl#############

from random import *
import openpyxl
characters =  string.digits
n_effettivo=0
riga = 1403
while riga < 1417:
        riga = riga + 1
        n_effettivo=int(n_effettivo)
        n_effettivo= n_effettivo+1
        password =  "".join(choice(characters) for x in range(randint(8, 8)))
        excel_document = openpyxl.load_workbook('listastudenti.xlsx')
        sheet = excel_document.get_sheet_by_name('Foglio1')
        nome_utente_cognome= str(sheet.cell(row = riga, column = 2).value)
        nome_utente_nome= str(sheet.cell(row = riga, column = 3).value)
        nome_utente= (nome_utente_nome+nome_utente_cognome)
        classe= str(sheet.cell(row = riga, column = 4).value)
        utente=str(nome_utente+classe)
        print (riga)
        print (utente)
        print (password)
        user=User.objects.create_user(utente, password=password)
        user.is_superuser=False
        user.is_staff=False
        user.save()


        n_effettivo= str(n_effettivo)
        pw = openpyxl.Workbook()
        coordinate = pw.get_sheet_by_name('Sheet')
        posizione= eval('coordinate["A'+n_effettivo+'"]')
        print(posizione)
        posizione_pass= eval('coordinate["B'+n_effettivo+'"]')
        posizione= nome_utente
        posizione_pass= password


pw.save('Pw.xlsx')



